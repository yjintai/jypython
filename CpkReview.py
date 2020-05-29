#!/usr/bin/env python
# -*- coding: utf-8 -*-
# CpkReview.py

import sys,getopt,os
import csv
import openpyxl
import datetime
import subprocess
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from win32com.client import Dispatch

VERSION = "R1.1"
CPK_SPEC_RED = 1.67
CPK_SPEC_YELLOW = 2.00

#从csv读取内容，转换成需要的excel格式.
def csv_to_xlsx(inputfile,outputfile):
    #打开csv文件
    try:
        if inputfile =="":
            csvfile = open('xmldata.csv','rb')
        else:
            csvfile = open(inputfile,'rb')
    except IOError:
        print("The csv file does not exist.")
        return
    #获取csv.reader
    csvlines = csv.reader(csvfile)
 
     #创建工作簿对象
    work_book = openpyxl.Workbook()
 
    #获取活动sheet
    work_sheet = work_book.active
    work_sheet.title = u"xmldata"
    #写入统计行的标题
    work_sheet.cell(row = 1, column = 3).value = "CPK"
    work_sheet.cell(row = 2, column = 3).value = "Cp"
    work_sheet.cell(row = 3, column = 3).value = "Ca"
    work_sheet.cell(row = 4, column = 3).value = "Counts"
    work_sheet.cell(row = 5, column = 3).value = "Average"
    work_sheet.cell(row = 6, column = 3).value = "SD"
    work_sheet.cell(row = 7, column = 3).value = "Min"
    work_sheet.cell(row = 8, column = 3).value = "Max"

    if outputfile =="":
        work_sheet.cell(row = 1, column = 1).value = "xmldata.xlsx"
    else:
        i=outputfile.rfind('\\');
        i=i-len(outputfile)+1;
        work_sheet.cell(row = 1, column = 1).value = outputfile[i:]
    
 
    #row
    row = 9
    count = 0
    #从第二行开始写入从csv读取的内容
    for line in csvlines:
        col = 1
        if 'Failed' in line:
            continue
        for cell in line:
            s = cell
            try:
                s = float(s)
            except ValueError:
                pass
            work_sheet.cell(row=row, column=col).value = s
            col += 1
        row += 1
        count += 1
    count -= 4
    #关闭CSV文件
    csvfile.close()

    #写入统计行的内容
    icolqty = 0
    for cell in work_sheet["13"]:
        if cell.value != "":
            icolqty += 1
    icolqty = work_sheet.max_column
    irowqty = work_sheet.max_row
    for icol in range(4,icolqty+1):    
        colletter = get_column_letter(icol)  
        work_sheet.cell(row = 5, column = icol).value = "=AVERAGE(xmldata!%s13:%s%s)"%(colletter,colletter,irowqty)
        work_sheet.cell(row = 6, column = icol).value = "=STDEV(xmldata!%s13:%s%s)"%(colletter,colletter,irowqty)
        work_sheet.cell(row = 7, column = icol).value = "=MIN(xmldata!%s13:%s%s)"%(colletter,colletter,irowqty)
        work_sheet.cell(row = 8, column = icol).value = "=MAX(xmldata!%s13:%s%s)"%(colletter,colletter,irowqty)
        work_sheet.cell(row = 4, column = icol).value = "=COUNT(xmldata!%s13:%s%s)"%(colletter,colletter,irowqty)
        work_sheet.cell(row = 3, column = icol).value = "=((%s5-((%s9+%s10)/2)))/((%s10-%s9)/2)"%(colletter,colletter,colletter,colletter,colletter)
        work_sheet.cell(row = 2, column = icol).value = "=(%s10-%s9)/(6*%s6)"%(colletter,colletter,colletter)
        work_sheet.cell(row = 1, column = icol).value = "=MIN((%s5-%s9)/(3*%s6),(%s10-%s5)/(3*%s6))"%(colletter,colletter,colletter,colletter,colletter,colletter)

    #保存工作表
    if outputfile =="":
        outputfile = "c:\\outputfile.xlsx"
    work_book.save(outputfile)
    work_book.close()

    #重新打开关闭EXCEL文件，以同时保存格式和数据，不执行这一步，以“data_only=True”打开是格式栏会只显示空数据
    just_open_excel(outputfile)

    work_book_dataonly = openpyxl.load_workbook(outputfile, data_only=True)
    ws_raw_dataonly =  work_book_dataonly["xmldata"]
    work_book = openpyxl.load_workbook(outputfile)
    ws_raw = work_book.active

    #Create Summary sheet.
    ws_Summary = work_book.create_sheet("Summary",0)

    #Create Failure Summary sheet.
    ws_Failure = work_book.create_sheet("Failure Summary",0)


    pf_Fill_Red = PatternFill("solid", fgColor="ff0000")
    pf_Fill_Yellow = PatternFill("solid", fgColor="ffff00")

    for icol in range(3,ws_raw.max_column+1):
        for irow in range(1,13):  
            if irow == 12: #项目名称移至第一列
                ws_Summary.cell(row = icol-2, column = 1).value = ws_raw.cell(irow,icol).value
            elif irow == 1 and icol > 3:
                ws_Summary.cell(row = icol-2, column = irow+1).value = "=MIN((F%d-J%d)/(3*G%d),(K%d-F%d)/(3*G%d))"%(icol-2,icol-2,icol-2,icol-2,icol-2,icol-2)
                if ws_raw_dataonly.cell(irow,icol).value < CPK_SPEC_RED:
                    ws_Summary.cell(row=icol-2, column=irow + 1).fill = pf_Fill_Red
                elif ws_raw_dataonly.cell(irow,icol).value < CPK_SPEC_YELLOW:
                    ws_Summary.cell(row=icol-2, column=irow + 1).fill = pf_Fill_Yellow
            elif irow == 2 and icol > 3:
                ws_Summary.cell(row = icol-2, column = irow+1).value = "=(K%d-J%d)/(6*G%d)"%(icol-2,icol-2,icol-2)
            elif irow == 3 and icol > 3:
                ws_Summary.cell(row = icol-2, column = irow+1).value = "=((F%d-((J%d+K%d)/2)))/((K%d-J%d)/2)"%(icol-2,icol-2,icol-2,icol-2,icol-2)
            else:
                ws_Summary.cell(row = icol-2, column = irow+1).value = ws_raw.cell(irow,icol).value
    ws_Summary.cell(row=1, column=1).value = "Test Item"


    irow_failure = 1
    for icol in range(3,ws_raw_dataonly.max_column+1):
        if icol ==3:
            ws_Failure.cell(row=irow_failure, column=1).value = "Test Item"
            ws_Failure.cell(row=irow_failure, column=2).value = "CPK"
            ws_Failure.cell(row=irow_failure, column=3).value = "Analysis"
            ws_Failure.cell(row=irow_failure, column=4).value = "Action"
            ws_Failure.cell(row=irow_failure, column=5).value = "Owner"
            irow_failure += 1
        elif ws_raw_dataonly.cell(1,icol).value < CPK_SPEC_YELLOW and ws_raw_dataonly.cell(12,icol).value not in{"G_Test_Duration","G_Test_Result","StationNO"}:
            ws_Failure.cell(row=irow_failure, column=1).value = ws_raw_dataonly.cell(12,icol).value
            ws_Failure.cell(row=irow_failure, column=2).value = ws_raw_dataonly.cell(1,icol).value
            if ws_raw_dataonly.cell(1,icol).value < CPK_SPEC_RED:
                ws_Failure.cell(row=irow_failure, column=2).fill = pf_Fill_Red
            else:
                ws_Failure.cell(row=irow_failure, column=2).fill = pf_Fill_Yellow
            irow_failure += 1
    ws_Summary.column_dimensions['A'].width = 50.0
    ws_Failure.column_dimensions['A'].width = 50.0
    ws_Failure.column_dimensions['C'].width = 30.0
    ws_Failure.column_dimensions['D'].width = 30.0
    ws_Failure.column_dimensions['E'].width = 20.0


    ws_Failure.auto_filter.ref = "A:B"
    ws_Failure.auto_filter.add_sort_condition("A:A")

    ws_Summary.auto_filter.ref = "A:L"
    ws_Summary.auto_filter.add_sort_condition("A:A")

    work_book.save(outputfile)
    work_book.close()



def just_open_excel(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()

#调用SEEKSTRING.EXE程序，从log中生成csv格式的报告
def callseekstring(productname,week,station,ccitem,limitfile,count):
    cmd = "SeekString.exe"
    parameters = " /T3 /X"+ccitem+" /C"+str(count)+" /IZ:\\rawdata\\"+productname+"\\"+week+"\\dbg "+"/S"+station+" /L"+limitfile
    cmd = cmd + parameters
    print(cmd)
    sub=subprocess.Popen(cmd,shell=True)
    sub.wait()
 
	
def getcurrentweek(wkcount):
    now_time = datetime.datetime.now()
    stryear = datetime.datetime.strftime(now_time,'%Y')
    strweek = datetime.datetime.strftime(now_time,'%W')
    iweek = int(strweek)
    yearweeklist=[]
    for i in range(0,wkcount):
        yearweeklist.append(stryear+"_"+str(iweek-i))
    return yearweeklist


def usage():
    print("\tFunction: Seek String from data server to csv file, \n\t\tand transfer to excel format with cpk calculation")
    print("\tpython.exe CpkReview.py [-h] [-p <productname>] [-s <station>] [-x <ccitem>] [-w <week>] [-c <count>]")
    print("\t[-h]: Display the help description.")
    print("\t[-p <productname>]: The product name, same format as \n\t\t\tone in the rawdata folder, eg: AR7586_UM...")
    print("\t[-s <station>]: The station name, same format as \n\t\t\tin the log name, eg: BFT,RFV,TTL...")
    print("\t[-x <ccitem>]: The item type, input for seekstring.exe,\n\t\t\t1:CC item, 2:Full item")
    print("\t[-w <week>]: The year week, same format as in subfolder \n\t\t\tof the productname folder, eg: 2018_41, \n\t\t\tor like 1,2,3...show the count of week from current week.")
    print("\t[-c <count>]: Count of log will be ramdonly selected for \n\t\t\tthe CPK calculation, 0 indicate all logs. ")

if __name__ == '__main__':
    print("Version: " + VERSION)
    opts, args = getopt.getopt(sys.argv[1:], "hp:s:x:w:c:")
    count = 0;
    for op, value in opts:
        if op == "-p":
            productname = value
        elif op == "-s":
            station = value
        elif op == "-w":
            wk = value
        elif op == "-x":
            ccitem = value
        elif op == "-c":
            count = value
        elif op == "-h":
            usage()
            sys.exit()
    
    yearweeklist=[]
    if len(wk) < 2:
        yearweeklist = getcurrentweek(int(wk))
    else:
        yearweeklist.append(wk)
    for week in yearweeklist:
        print ("Current Week: "+week)
        input_file = "Z:\\rawdata\\"+productname+"\\"+week+"\\dbg\\xmldata.csv"
        output_file = "Z:\\report\\CpkReport\\"+productname+"_"+week+"_"+station+".xlsx"
        if "AR758" in productname:
            limit_file = "Z:\\report\\CpkReport\\limit\\AR758X\\"+station+"_Limits_AR758x.xml"
        elif "AR759" in productname:
            limit_file = "Z:\\report\\CpkReport\\limit\\AR759X\\"+station+"_Limits_AR759x.xml"
        elif "AR755" in productname or "AR865" in productname:
            limit_file = "Z:\\report\\CpkReport\\limit\\AR755X\\"+station+"_Limits_AR755x.xml" 
        else:
            limit_file = "N"
        print limit_file
        if 1 > 2:
            """os.access(output_file, os.F_OK):"""
            print (output_file + " already exist!")
        else:
            print("Call SeekString app to create .csv report: "+productname+", "+station +", "+week)
            callseekstring(productname,week,station,ccitem,limit_file,count)

            print("transform csv file to xlsx format")
            csv_to_xlsx(input_file, output_file)
